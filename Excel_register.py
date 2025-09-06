import openpyxl

def load_data():
    excel_path = 'D:/Poornima/Ofc/Automation_class/Code/Selenium/Register_multiple_user/test_data.xlsx'
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook['test_data']
    max_row_ = sheet.max_row #2
    max_col_ = sheet.max_column #2
    print(f"Maximum row count {max_row_} and maximimum column count is {max_col_}")
    data_list = [] # []
    for row in range(2,6):
            firstname = sheet.cell(row=row,column=1).value
            lastname  = sheet.cell(row=row,column=2).value
            username  = sheet.cell(row=row,column=3).value
            password  = sheet.cell(row=row,column=4).value
            data_list.append((firstname,lastname,username,password))
    return data_list
